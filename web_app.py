import os
import io
import re
import tempfile
from datetime import datetime, timedelta, date, time as time_obj
from urllib.parse import urlparse

import pandas as pd
import psycopg2
import requests
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, abort, after_this_request
from PIL import Image, ImageDraw, ImageFont
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL")
SECRET_KEY = os.getenv("WEB_SECRET_KEY", "change-me")
TIMEZONE = os.getenv("TIMEZONE", "Asia/Tashkent")
STORE_LOCATION_NAME = os.getenv("STORE_LOCATION_NAME", "Do'kon")
SELLER_PHONE = os.getenv("SELLER_PHONE", "+998330131992")
SELLER_NAME = os.getenv("SELLER_NAME", "")
ADMIN_TELEGRAM_IDS = os.getenv("ADMIN_TELEGRAM_IDS", "")

DEFAULT_ADMIN_IDS = {1262207928, 963690743, 8450201406}

app = Flask(__name__)
app.secret_key = SECRET_KEY

USD_RATE_CACHE = {"rate": None, "time": None}
CYRILLIC_PATTERN = re.compile(r'[А-Яа-яЁёҢғқўҳ]', flags=re.UNICODE)


def parse_admin_ids(raw_value: str):
    if not raw_value:
        return set(DEFAULT_ADMIN_IDS)
    values = set()
    for chunk in raw_value.replace(";", ",").split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        try:
            values.add(int(chunk))
        except ValueError:
            continue
    return values or set(DEFAULT_ADMIN_IDS)


ADMIN_IDS = parse_admin_ids(ADMIN_TELEGRAM_IDS)


if not DATABASE_URL:
    raise SystemExit("DATABASE_URL ni .env ga qo'ying")


def get_conn():
    url = urlparse(DATABASE_URL)
    return psycopg2.connect(
        dbname=url.path[1:],
        user=url.username,
        password=url.password,
        host=url.hostname,
        port=url.port
    )


def init_db():
    conn = get_conn()
    cur = conn.cursor()
    raw_sql = open("db_init.sql", "r", encoding="utf-8").read().splitlines()
    filtered = []
    for line in raw_sql:
        stripped = line.strip()
        if stripped.startswith(("@@", "diff --git", "---", "+++", "index ")):
            continue
        filtered.append(line)
    sql = "\n".join(filtered)
    cur.execute(sql)
    conn.commit()
    cur.close()
    conn.close()


def format_money(value):
    try:
        return f"{int(value):,}".replace(",", ".") + " so'm"
    except Exception:
        return str(value)


def contains_cyrillic(text: str):
    if not isinstance(text, str):
        return False
    return bool(CYRILLIC_PATTERN.search(text))


def get_usd_rate():
    global USD_RATE_CACHE
    now = datetime.utcnow()
    if USD_RATE_CACHE["rate"] and USD_RATE_CACHE["time"] and now - USD_RATE_CACHE["time"] < timedelta(hours=24):
        return USD_RATE_CACHE["rate"]

    try:
        resp = requests.get("https://cbu.uz/uz/arkhiv-kursov-valyut/json/USD/")
        data = resp.json()
        if isinstance(data, list) and len(data) > 0:
            rate = float(data[0]["Rate"])
            USD_RATE_CACHE = {"rate": rate, "time": now}
            return rate
    except Exception:
        pass

    return USD_RATE_CACHE["rate"] or 12800.0


def _get_font(size=16):
    candidates = [
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _measure_text(draw, text, font):
    try:
        bbox = draw.textbbox((0, 0), text, font=font)
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        return (w, h)
    except Exception:
        pass

    try:
        size = draw.textsize(text, font=font)
        return (size[0], size[1])
    except Exception:
        pass

    try:
        bbox = font.getbbox(text)
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        return (w, h)
    except Exception:
        pass

    try:
        size = font.getsize(text)
        return (size[0], size[1])
    except Exception:
        pass

    approx_w = int(len(text) * (getattr(font, "size", 12) * 0.6))
    approx_h = int((getattr(font, "size", 12)) * 1.2)
    return (approx_w, approx_h)


def receipt_text(sale_id):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """
        SELECT s.id, s.total_amount, s.payment_type, s.created_at,
               c.name as cust_name, c.phone as cust_phone
        FROM sales s
        LEFT JOIN customers c ON s.customer_id = c.id
        WHERE s.id=%s;
        """,
        (sale_id,),
    )
    sale = cur.fetchone()
    cur.execute("SELECT name, qty, price, total FROM sale_items WHERE sale_id=%s;", (sale_id,))
    items = cur.fetchall()
    cur.close()
    conn.close()

    if not sale:
        return "Sotuv topilmadi."

    created_at = sale.get("created_at")
    if isinstance(created_at, datetime):
        created_at = created_at + timedelta(hours=5)

    seller_display = f"{SELLER_NAME} {SELLER_PHONE}" if SELLER_NAME else f"{SELLER_PHONE}"

    lines = [
        f"🧾 Chek №{sale_id}",
        f"📅 Sana: {created_at.strftime('%d.%m.%Y %H:%M:%S') if created_at else ''}",
        f"🏬 Do‘kon: {STORE_LOCATION_NAME}",
        f"👨‍💼 Sotuvchi: {seller_display}",
        f"👤 Mijoz: {sale.get('cust_name') or '-'} {sale.get('cust_phone') or ''}",
        "────────────────────────────",
    ]
    for it in items:
        lines.append(
            f"{it.get('name')} — {it.get('qty')} x {format_money(it.get('price'))} = {format_money(it.get('total'))}"
        )
    lines.extend(
        [
            "────────────────────────────",
            f"💰 Jami: {format_money(sale.get('total_amount') or 0)}",
            f"💳 To‘lov turi: {sale.get('payment_type')}",
            "────────────────────────────",
            "Tashrifingiz uchun rahmat! ❤️",
        ]
    )
    return "\n".join(lines)


def format_som_plain(v: int) -> str:
    try:
        return f"{int(v):,}".replace(",", " ")
    except Exception:
        return str(v)

def _wrap_text(draw, text, font, max_width):
    words = (text or "").split()
    if not words:
        return [""]
    lines, cur = [], ""
    for w in words:
        test = (cur + " " + w).strip()
        tw, _ = _measure_text(draw, test, font)
        if tw <= max_width:
            cur = test
        else:
            if cur:
                lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines

def receipt_image_bytes(sale_id):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT s.id, s.total_amount, s.payment_type, s.created_at,
               c.name AS cust_name, c.phone AS cust_phone
        FROM sales s
        LEFT JOIN customers c ON s.customer_id = c.id
        WHERE s.id = %s;
    """, (sale_id,))
    sale = cur.fetchone()

    cur.execute("""
        SELECT name, qty, price, total
        FROM sale_items
        WHERE sale_id = %s
        ORDER BY id;
    """, (sale_id,))
    items = cur.fetchall()

    cur.close()
    conn.close()

    if not sale:
        return None

    created = sale.get("created_at")
    if isinstance(created, datetime):
        created_local = created + timedelta(hours=5)
    else:
        created_local = datetime.utcnow() + timedelta(hours=5)

    # 80mm thermal: 576px width (printer-friendly)
    W = 576
    P = 22
    GAP = 8

    font_brand = _get_font(34)
    font_title = _get_font(26)
    font_bold = _get_font(22)
    font = _get_font(20)
    font_small = _get_font(18)

    temp = Image.new("RGB", (W, 10), "white")
    d = ImageDraw.Draw(temp)

    # columns (ITEM | QTY | PRICE | TOTAL)
    col_name_w = W - (P * 2) - 240
    col_qty_w = 60
    col_price_w = 90
    col_total_w = 90

    seller_display = f"{SELLER_NAME} ({SELLER_PHONE})" if SELLER_NAME else f"{SELLER_PHONE}"
    cust_line = f"{sale.get('cust_name') or '-'} {sale.get('cust_phone') or ''}".strip()

    total_amount = int(sale.get("total_amount") or 0)
    pay_type = (sale.get("payment_type") or "-").upper()

    blocks = []
    blocks.append(("center", "SRM", font_brand))
    blocks.append(("center", "SALES RECEIPT", font_title))
    blocks.append(("hr", "", None))

    blocks.append(("kv", ("Chek ID", f"#{sale_id}", font)))
    blocks.append(("kv", ("Sana", created_local.strftime("%d.%m.%Y %H:%M"), font)))
    blocks.append(("kv", ("To'lov", pay_type, font_bold)))
    blocks.append(("kv", ("Sotuvchi", seller_display, font_small)))
    blocks.append(("kv", ("Mijoz", cust_line, font_small)))

    blocks.append(("hr", "", None))
    blocks.append(("table_head", "", None))

    for it in items:
        name = str(it.get("name") or "").strip()
        qty = int(it.get("qty") or 0)
        price = int(it.get("price") or 0)
        total = int(it.get("total") or (qty * price))

        name_lines = _wrap_text(d, name, font, col_name_w)
        blocks.append(("row", {
            "name": name_lines[0],
            "qty": str(qty),
            "price": format_som_plain(price),
            "total": format_som_plain(total),
            "font": font
        }))
        for extra in name_lines[1:]:
            blocks.append(("row_sub", {"name": extra, "font": font}))

    blocks.append(("hr", "", None))
    blocks.append(("sum", ("JAMI", f"{format_som_plain(total_amount)} so'm", font_brand)))
    blocks.append(("hr", "", None))
    blocks.append(("center_small", "Tashrifingiz uchun rahmat!", font_small))

    # ---- height calc ----
    tmp = Image.new("RGB", (W, 10), "white")
    draw_tmp = ImageDraw.Draw(tmp)

    H = P
    def add_h(text, fnt, extra=GAP):
        nonlocal H
        _, hh = _measure_text(draw_tmp, text, fnt)
        H += hh + extra

    for kind, payload, fnt in blocks:
        if kind in ("center", "center_small"):
            add_h(payload, fnt, GAP)
        elif kind == "hr":
            H += 18
        elif kind == "kv":
            k, v, ff = payload
            add_h(f"{k}: {v}", ff, 6)
        elif kind == "table_head":
            H += 36
        elif kind in ("row", "row_sub"):
            add_h(payload["name"], payload.get("font", font), 6)
        elif kind == "sum":
            k, v, ff = payload
            add_h(f"{k} {v}", ff, 10)

    qr_size = 180
    H += qr_size + 30 + P
    H = max(720, H)

    img = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)

    y = P

    def hr():
        nonlocal y
        y += 6
        draw.line((P, y, W - P, y), fill=(0, 0, 0), width=2)
        y += 12

    def center(text, fnt):
        nonlocal y
        tw, th = _measure_text(draw, text, fnt)
        draw.text(((W - tw) // 2, y), text, font=fnt, fill="black")
        y += th + GAP

    def kv(k, v, fnt):
        nonlocal y
        left = f"{k}:"
        draw.text((P, y), left, font=fnt, fill="black")
        vw, vh = _measure_text(draw, str(v), fnt)
        draw.text((W - P - vw, y), str(v), font=fnt, fill="black")
        _, lh = _measure_text(draw, left, fnt)
        y += max(lh, vh) + 6

    def table_head():
        nonlocal y
        draw.text((P, y), "ITEM", font=font_bold, fill="black")
        draw.text((P + col_name_w + 10, y), "QTY", font=font_bold, fill="black")
        draw.text((P + col_name_w + 10 + col_qty_w, y), "PRICE", font=font_bold, fill="black")
        draw.text((W - P - col_total_w + 10, y), "TOTAL", font=font_bold, fill="black")
        y += 26
        draw.line((P, y, W - P, y), fill=(0, 0, 0), width=1)
        y += 10

    def row(name, qty=None, price=None, total=None, fnt=font):
        nonlocal y
        draw.text((P, y), name, font=fnt, fill="black")

        if qty is not None:
            qw, _ = _measure_text(draw, str(qty), fnt)
            draw.text((P + col_name_w + 10 + (col_qty_w - qw) // 2, y), str(qty), font=fnt, fill="black")

        if price is not None:
            pw, _ = _measure_text(draw, str(price), fnt)
            draw.text((P + col_name_w + 10 + col_qty_w + (col_price_w - pw), y), str(price), font=fnt, fill="black")

        if total is not None:
            tw, _ = _measure_text(draw, str(total), fnt)
            draw.text((W - P - tw, y), str(total), font=fnt, fill="black")

        _, nh = _measure_text(draw, name, fnt)
        y += nh + 6

    def sum_line(label, value, fnt):
        nonlocal y
        draw.text((P, y), label, font=fnt, fill="black")
        vw, vh = _measure_text(draw, value, fnt)
        draw.text((W - P - vw, y), value, font=fnt, fill="black")
        _, lh = _measure_text(draw, label, fnt)
        y += max(lh, vh) + 10

    for kind, payload, fnt in blocks:
        if kind == "center":
            center(payload, fnt)
        elif kind == "center_small":
            center(payload, fnt)
        elif kind == "hr":
            hr()
        elif kind == "kv":
            k, v, ff = payload
            kv(k, v, ff)
        elif kind == "table_head":
            table_head()
        elif kind == "row":
            row(payload["name"], payload["qty"], payload["price"], payload["total"], payload.get("font", font))
        elif kind == "row_sub":
            row(payload["name"], None, None, None, payload.get("font", font))
        elif kind == "sum":
            k, v, ff = payload
            sum_line(k, v, ff)

    # QR
    try:
        import qrcode
        qr_payload = f"SRM|sale:{sale_id}|total:{total_amount}|time:{created_local.strftime('%Y-%m-%d %H:%M')}"
        qr = qrcode.make(qr_payload).resize((qr_size, qr_size))
        img.paste(qr, ((W - qr_size) // 2, H - qr_size - P - 10))
    except Exception:
        pass

    buf = io.BytesIO()
    buf.name = f"receipt_{sale_id}.png"
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def login_required(role=None):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = session.get("user")
            if not user:
                return redirect(url_for("login"))
            if role and user.get("role") != role:
                abort(403)
            return fn(*args, **kwargs)

        return wrapper

    return decorator


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        login_type = request.form.get("login_type")
        if login_type == "admin":
            telegram_id = request.form.get("telegram_id", "").strip()
            if telegram_id.isdigit() and int(telegram_id) in ADMIN_IDS:
                session["user"] = {"role": "admin", "username": f"admin-{telegram_id}"}
                return redirect(url_for("dashboard"))
            flash("Admin telegram ID noto'g'ri.", "error")
        else:
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()
            conn = get_conn()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute(
                "SELECT username, password_hash, role, is_active FROM web_users WHERE username=%s;",
                (username,),
            )
            row = cur.fetchone()
            cur.close()
            conn.close()
            if row and row.get("is_active") and check_password_hash(row["password_hash"], password):
                session["user"] = {"role": row["role"], "username": row["username"]}
                return redirect(url_for("dashboard"))
            flash("Login yoki parol noto'g'ri.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))


@app.route("/")
@login_required()
def dashboard():
    return render_template("dashboard.html", user=session.get("user"))


@app.route("/admin/users", methods=["GET", "POST"])
@login_required(role="admin")
def admin_users():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if not username or not password:
            flash("Login va parol kiriting.", "error")
        else:
            password_hash = generate_password_hash(password)
            conn = get_conn()
            cur = conn.cursor()
            try:
                cur.execute(
                    "INSERT INTO web_users (username, password_hash, role) VALUES (%s, %s, 'seller');",
                    (username, password_hash),
                )
                conn.commit()
                flash("Sotuvchi yaratildi.", "success")
            except psycopg2.Error:
                conn.rollback()
                flash("Bu login band.", "error")
            finally:
                cur.close()
                conn.close()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT id, username, role, is_active, created_at FROM web_users ORDER BY id;")
    users = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("admin_users.html", users=users)


@app.route("/products")
@login_required()
def products():
    search = request.args.get("q", "").strip()
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if search:
        cur.execute(
            "SELECT * FROM products WHERE name ILIKE %s ORDER BY id;",
            (f"%{search}%",),
        )
    else:
        cur.execute("SELECT * FROM products ORDER BY id;")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("products.html", products=rows, search=search)


@app.route("/products/add", methods=["GET", "POST"])
@login_required()
def products_add():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        qty = request.form.get("qty", "0").strip()
        cost_price_usd = request.form.get("cost_price_usd", "0").strip()
        suggest_price = request.form.get("suggest_price", "0").strip()

        if not name:
            flash("Mahsulot nomi kerak.", "error")
            return redirect(url_for("products_add"))
        if contains_cyrillic(name):
            flash("Iltimos, lotincha yozing.", "error")
            return redirect(url_for("products_add"))

        try:
            qty_val = int(qty)
            cost_usd_val = float(cost_price_usd.replace(",", "."))
            suggest_val = int(suggest_price.replace(" ", ""))
        except ValueError:
            flash("Miqdor va narxlar raqam bo'lishi kerak.", "error")
            return redirect(url_for("products_add"))

        usd_rate = get_usd_rate()
        cost_som = int(cost_usd_val * usd_rate)

        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT id, qty FROM products WHERE name = %s AND cost_price_usd = %s;",
            (name, cost_usd_val),
        )
        existing = cur.fetchone()
        if existing:
            new_qty = existing[1] + qty_val
            cur.execute(
                "UPDATE products SET qty=%s, cost_price=%s, usd_rate=%s, suggest_price=%s WHERE id=%s;",
                (new_qty, cost_som, usd_rate, suggest_val, existing[0]),
            )
        else:
            cur.execute(
                """
                INSERT INTO products (name, qty, cost_price, cost_price_usd, usd_rate, suggest_price)
                VALUES (%s, %s, %s, %s, %s, %s);
                """,
                (name, qty_val, cost_som, cost_usd_val, usd_rate, suggest_val),
            )
        conn.commit()
        cur.close()
        conn.close()
        flash("Mahsulot saqlandi.", "success")
        return redirect(url_for("products"))

    return render_template("products_add.html")


@app.route("/products/upload", methods=["GET", "POST"])
@login_required()
def products_upload():
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            flash("Excel fayl tanlang.", "error")
            return redirect(url_for("products_upload"))
        try:
            df = pd.read_excel(file, engine="openpyxl")
        except Exception:
            df = pd.read_excel(file)

        df.columns = [c.strip().lower() for c in df.columns]
        name_keys = ["name", "nom", "product", "product_name", "mahsulot", "mahsol", "mahsulot nomi"]
        qty_keys = ["qty", "quantity", "soni", "miqdor", "son"]
        cost_usd_keys = ["cost_price_usd", "cost_usd", "opt_narx_usd", "usd narx"]
        suggest_keys = ["suggest_price", "sell_price", "price", "sotuv_narx", "taklif narxi"]

        def find_col(keys):
            for key in keys:
                if key in df.columns:
                    return key
            return None

        col_name = find_col(name_keys)
        col_qty = find_col(qty_keys)
        col_cost_usd = find_col(cost_usd_keys)
        col_suggest = find_col(suggest_keys)

        if not col_name or not col_qty or not col_cost_usd:
            flash("Excel faylda nom, miqdor yoki USD narx ustunlari topilmadi.", "error")
            return redirect(url_for("products_upload"))

        df = df[[col_name, col_qty, col_cost_usd] + ([col_suggest] if col_suggest else [])].copy()
        df[col_name] = df[col_name].astype(str).str.strip()
        df[col_qty] = df[col_qty].apply(lambda x: int(float(str(x).replace(",", "").strip())) if pd.notna(x) else 0)
        df[col_cost_usd] = df[col_cost_usd].apply(lambda x: float(str(x).replace(",", ".").strip()) if pd.notna(x) else 0)
        if col_suggest:
            df[col_suggest] = df[col_suggest].apply(lambda x: int(float(str(x).replace(",", "").strip())) if pd.notna(x) else 0)
        else:
            df["suggest_temp"] = 0
            col_suggest = "suggest_temp"

        conn = get_conn()
        cur = conn.cursor()
        inserted = 0
        updated = 0
        skipped = 0
        for _, row in df.iterrows():
            pname = str(row[col_name]).strip()
            pqty = int(row[col_qty])
            pcost_usd = float(row[col_cost_usd])
            psuggest = int(row[col_suggest])
            usd_rate = get_usd_rate()
            pcost_som = int(pcost_usd * usd_rate)

            if not pname or pqty <= 0:
                skipped += 1
                continue

            cur.execute(
                "SELECT id FROM products WHERE name ILIKE %s AND cost_price_usd = %s LIMIT 1;",
                (pname, pcost_usd),
            )
            existing = cur.fetchone()
            if existing:
                cur.execute(
                    """
                    UPDATE products
                    SET qty = qty + %s,
                        cost_price = %s,
                        usd_rate = %s,
                        suggest_price = COALESCE(%s, suggest_price)
                    WHERE id=%s;
                    """,
                    (pqty, pcost_som, usd_rate, psuggest if psuggest > 0 else None, existing[0]),
                )
                updated += 1
            else:
                cur.execute(
                    """
                    INSERT INTO products (name, qty, cost_price, cost_price_usd, usd_rate, suggest_price)
                    VALUES (%s, %s, %s, %s, %s, %s);
                    """,
                    (pname, pqty, pcost_som, pcost_usd, usd_rate, psuggest if psuggest > 0 else None),
                )
                inserted += 1

        conn.commit()
        cur.close()
        conn.close()
        flash(f"Excel yuklandi. Yangi: {inserted}, yangilangan: {updated}, o'tkazib yuborilgan: {skipped}.", "success")
        return redirect(url_for("products"))

    return render_template("products_upload.html")


def get_cart():
    return session.setdefault("cart", [])


def clear_cart():
    session["cart"] = []


@app.route("/sales/new")
@login_required()
def sales_new():
    search = request.args.get("q", "").strip()
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    if search:
        cur.execute(
            "SELECT id, name, qty, suggest_price FROM products WHERE name ILIKE %s AND qty > 0 ORDER BY id;",
            (f"%{search}%",),
        )
    else:
        cur.execute("SELECT id, name, qty, suggest_price FROM products WHERE qty > 0 ORDER BY id;")
    products_list = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("sales_new.html", products=products_list, search=search)


@app.route("/sales/cart", methods=["GET", "POST"])
@login_required()
def sales_cart():
    if request.method == "POST":
        product_id = request.form.get("product_id")
        qty = request.form.get("qty")
        price = request.form.get("price")
        if not (product_id and qty and price):
            flash("Mahsulot, miqdor va narx kerak.", "error")
            return redirect(url_for("sales_new"))
        try:
            product_id = int(product_id)
            qty = int(qty)
            price = int(price)
        except ValueError:
            flash("Miqdor va narx raqam bo'lishi kerak.", "error")
            return redirect(url_for("sales_new"))

        conn = get_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT id, name, qty, suggest_price FROM products WHERE id=%s;", (product_id,))
        product = cur.fetchone()
        cur.close()
        conn.close()
        if not product:
            flash("Mahsulot topilmadi.", "error")
            return redirect(url_for("sales_new"))
        if qty > product["qty"]:
            flash("Omborda yetarli miqdor yo'q.", "error")
            return redirect(url_for("sales_new"))

        cart = get_cart()
        cart.append({
            "product_id": product["id"],
            "name": product["name"],
            "qty": qty,
            "price": price,
        })
        session.modified = True
        flash("Savatchaga qo'shildi.", "success")
        return redirect(url_for("sales_cart"))

    cart = get_cart()
    total = sum(item["qty"] * item["price"] for item in cart)
    return render_template("sales_cart.html", cart=cart, total=total)


@app.route("/sales/cart/remove/<int:index>")
@login_required()
def sales_cart_remove(index):
    cart = get_cart()
    if 0 <= index < len(cart):
        cart.pop(index)
        session.modified = True
    return redirect(url_for("sales_cart"))


@app.route("/sales/checkout", methods=["GET", "POST"])
@login_required()
def sales_checkout():
    cart = get_cart()
    if not cart:
        flash("Savatcha bo'sh.", "error")
        return redirect(url_for("sales_new"))

    if request.method == "POST":
        customer_type = request.form.get("customer_type")
        payment_type = request.form.get("payment_type")
        if payment_type not in {"naqd", "qarz"}:
            flash("To'lov turini tanlang.", "error")
            return redirect(url_for("sales_checkout"))

        conn = get_conn()
        cur = conn.cursor()
        try:
            if customer_type == "new":
                name = request.form.get("customer_name", "").strip()
                phone = request.form.get("customer_phone", "").strip()
                cur.execute("INSERT INTO customers (name, phone) VALUES (%s, %s) RETURNING id;", (name, phone))
                customer_id = cur.fetchone()[0]
            else:
                customer_id = int(request.form.get("customer_id", "0"))

            total_amount = sum(item["qty"] * item["price"] for item in cart)
            cur.execute(
                """
                INSERT INTO sales (customer_id, total_amount, payment_type, seller_phone)
                VALUES (%s, %s, %s, %s)
                RETURNING id;
                """,
                (customer_id, total_amount, payment_type, SELLER_PHONE),
            )
            sale_id = cur.fetchone()[0]

            for item in cart:
                cur.execute(
                    """
                    INSERT INTO sale_items (sale_id, product_id, name, qty, price, total)
                    VALUES (%s, %s, %s, %s, %s, %s);
                    """,
                    (sale_id, item["product_id"], item["name"], item["qty"], item["price"], item["qty"] * item["price"]),
                )
                cur.execute(
                    "UPDATE products SET qty = qty - %s WHERE id=%s;",
                    (item["qty"], item["product_id"]),
                )

            if payment_type == "qarz":
                cur.execute(
                    "INSERT INTO debts (customer_id, sale_id, amount) VALUES (%s, %s, %s);",
                    (customer_id, sale_id, total_amount),
                )

            conn.commit()
        except Exception:
            conn.rollback()
            flash("Savdoni saqlashda xatolik.", "error")
            return redirect(url_for("sales_checkout"))
        finally:
            cur.close()
            conn.close()

        clear_cart()
        return redirect(url_for("sales_receipt", sale_id=sale_id))

    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT id, name, phone FROM customers ORDER BY id DESC LIMIT 50;")
    customers = cur.fetchall()
    cur.close()
    conn.close()

    total = sum(item["qty"] * item["price"] for item in cart)
    return render_template("sales_checkout.html", cart=cart, total=total, customers=customers)


@app.route("/sales/receipt/<int:sale_id>")
@login_required()
def sales_receipt(sale_id):
    text = receipt_text(sale_id)
    return render_template("sales_receipt.html", sale_id=sale_id, receipt_text=text)


@app.route("/sales/receipt/<int:sale_id>/image")
@login_required()
def sales_receipt_image(sale_id):
    buf = receipt_image_bytes(sale_id)
    if not buf:
        abort(404)
    return send_file(buf, mimetype="image/png", download_name=f"receipt_{sale_id}.png")


@app.route("/stats")
@login_required()
def stats_home():
    return render_template("stats.html")


def period_range(period_key):
    now = datetime.utcnow() + timedelta(hours=5)
    today = now.date()
    if period_key == "daily":
        start = datetime.combine(today, time_obj.min)
        end = start + timedelta(days=1)
    elif period_key == "monthly":
        start = datetime.combine(date(today.year, today.month, 1), time_obj.min)
        if today.month == 12:
            end = datetime.combine(date(today.year + 1, 1, 1), time_obj.min)
        else:
            end = datetime.combine(date(today.year, today.month + 1, 1), time_obj.min)
    elif period_key == "yearly":
        start = datetime.combine(date(today.year, 1, 1), time_obj.min)
        end = datetime.combine(date(today.year + 1, 1, 1), time_obj.min)
    else:
        raise ValueError("Unknown period")
    return start, end


def generate_stats_df(start_dt, end_dt):
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """
        SELECT si.product_id, si.name AS product_name,
               SUM(si.qty) AS sold_qty,
               SUM(si.total) AS total_sold,
               COALESCE(p.cost_price,0) AS cost_price
        FROM sale_items si
        JOIN sales s ON s.id = si.sale_id
        LEFT JOIN products p ON p.id = si.product_id
        WHERE s.created_at >= %s AND s.created_at < %s
        GROUP BY si.product_id, si.name, p.cost_price
        ORDER BY si.name;
        """,
        (start_dt, end_dt),
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        return pd.DataFrame(columns=["product_id", "name", "sold_qty", "cost_price", "total_sold", "total_cost", "profit"])

    df = pd.DataFrame(rows)
    df = df.rename(columns={"product_name": "name"})
    df["sold_qty"] = df["sold_qty"].astype(int)
    df["total_sold"] = df["total_sold"].astype(int)
    df["cost_price"] = df["cost_price"].astype(int)
    df["total_cost"] = df["sold_qty"] * df["cost_price"]
    df["profit"] = df["total_sold"] - df["total_cost"]
    df = df[["product_id", "name", "sold_qty", "cost_price", "total_sold", "total_cost", "profit"]]
    return df


def make_excel_from_df(df, title, start_dt, end_dt):
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        meta = pd.DataFrame([
            {
                "Hisobot": title,
                "Sana boshi": start_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "Sana oxiri": end_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "Yaratildi": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        ])
        meta.to_excel(writer, index=False, sheet_name="Meta")
        if df.empty:
            pd.DataFrame([{ "Xabar": "Ushbu davrda hech qanday mahsulot sotilmagan." }]).to_excel(
                writer, index=False, sheet_name="Hisobot"
            )
        else:
            df.to_excel(writer, index=False, sheet_name="Hisobot")
            ws = writer.sheets["Hisobot"]
            start_row = len(df) + 2
            ws.cell(row=start_row, column=2, value="Jami")
            ws.cell(row=start_row, column=3, value=int(df["sold_qty"].sum()))
            ws.cell(row=start_row, column=5, value=int(df["total_sold"].sum()))
            ws.cell(row=start_row, column=6, value=int(df["total_cost"].sum()))
            ws.cell(row=start_row, column=7, value=int(df["profit"].sum()))

    out.seek(0)
    return out


@app.route("/stats/report/<period>")
@login_required()
def stats_report(period):
    try:
        start_dt, end_dt = period_range(period)
    except ValueError:
        abort(404)
    df = generate_stats_df(start_dt, end_dt)
    title = f"{period.title()} hisobot"
    excel_buf = make_excel_from_df(df, title, start_dt, end_dt)
    filename = f"hisobot_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(excel_buf, as_attachment=True, download_name=filename)


@app.route("/stats/sale", methods=["POST"])
@login_required()
def stats_sale_report():
    sale_id = request.form.get("sale_id", "").strip()
    if not sale_id.isdigit():
        flash("Sotuv ID noto'g'ri.", "error")
        return redirect(url_for("stats_home"))

    sale_id = int(sale_id)
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """
        SELECT s.id AS sale_id, s.created_at, s.total_amount, s.payment_type, c.name as cust_name, c.phone as cust_phone
        FROM sales s
        LEFT JOIN customers c ON c.id = s.customer_id
        WHERE s.id = %s;
        """,
        (sale_id,),
    )
    sale = cur.fetchone()
    if not sale:
        cur.close()
        conn.close()
        flash("Sotuv topilmadi.", "error")
        return redirect(url_for("stats_home"))

    cur.execute(
        """
        SELECT si.product_id, si.name, si.qty, si.price, si.total, COALESCE(p.cost_price,0) AS cost_price
        FROM sale_items si
        LEFT JOIN products p ON p.id = si.product_id
        WHERE si.sale_id = %s;
        """,
        (sale_id,),
    )
    items = cur.fetchall()
    cur.close()
    conn.close()

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        sale_meta = pd.DataFrame([
            {
                "Sale ID": sale["sale_id"],
                "Sana": sale["created_at"].strftime("%Y-%m-%d %H:%M:%S") if sale["created_at"] else "",
                "Mijoz": sale.get("cust_name") or "",
                "Telefon": sale.get("cust_phone") or "",
                "To'lov turi": sale.get("payment_type") or "",
                "Jami summa": sale.get("total_amount") or 0,
            }
        ])
        sale_meta.to_excel(writer, index=False, sheet_name="Sale")
        if not items:
            pd.DataFrame([{ "Xabar": "Ushbu chekda elementlar yo'q" }]).to_excel(
                writer, index=False, sheet_name="Items"
            )
        else:
            df_items = pd.DataFrame(items)
            df_items.to_excel(writer, index=False, sheet_name="Items")
    out.seek(0)
    filename = f"chek_{sale_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(out, as_attachment=True, download_name=filename)


@app.route("/debts")
@login_required()
def debts():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """
        SELECT d.id, d.amount, d.created_at, c.name, c.phone
        FROM debts d
        JOIN customers c ON d.customer_id = c.id
        ORDER BY d.created_at DESC;
        """
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("debts.html", debts=rows, format_money=format_money)


@app.route("/debts/export")
@login_required()
def debts_export():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """
        SELECT c.name AS Mijoz, c.phone AS Telefon, d.amount AS Qarz_summasi, d.created_at AS Sana
        FROM debts d
        JOIN customers c ON d.customer_id = c.id
        ORDER BY d.created_at DESC;
        """
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        flash("Qarzdorlar topilmadi.", "error")
        return redirect(url_for("debts"))

    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Qarzdorlar")
    buf.seek(0)
    filename = f"qarzdorlar_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename)


@app.route("/stock/export")
@login_required()
def stock_export():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """
        SELECT id, name, qty, cost_price_usd, cost_price, suggest_price, created_at
        FROM products ORDER BY id;
        """
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        flash("Omborda mahsulot yo'q.", "error")
        return redirect(url_for("products"))

    df = pd.DataFrame(rows)
    df.rename(
        columns={
            "id": "№",
            "name": "Mahsulot nomi",
            "qty": "Miqdor (dona)",
            "cost_price_usd": "Narx (USD)",
            "cost_price": "Narx (so‘m)",
            "suggest_price": "Taklif narxi (so‘m)",
            "created_at": "Qo‘shilgan sana",
        },
        inplace=True,
    )

    total_qty = df["Miqdor (dona)"].sum()
    total_usd = df["Narx (USD)"].sum()
    total_som = df["Narx (so‘m)"].sum()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file_path = tmp.name

    df.to_excel(file_path, index=False, sheet_name="Ombor")

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(file_path)
    ws = wb["Ombor"]
    last_row = ws.max_row + 2

    ws.cell(row=last_row, column=2, value="Jami:")
    ws.cell(row=last_row, column=3, value=total_qty)
    ws.cell(row=last_row, column=4, value=total_usd)
    ws.cell(row=last_row, column=5, value=total_som)

    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    for col in range(2, 6):
        cell = ws.cell(row=last_row, column=col)
        cell.font = bold_font
        cell.fill = yellow_fill

    wb.save(file_path)

    filename = f"ombor_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    @after_this_request
    def cleanup(response):
        try:
            os.remove(file_path)
        except OSError:
            pass
        return response

    return send_file(file_path, as_attachment=True, download_name=filename)


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "8000")), debug=True)
